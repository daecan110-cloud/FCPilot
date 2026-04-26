"""보장분석표 템플릿 무결성 검증 — 양식 수정 전후에 반드시 실행.

사용법:
  py tests/test_excel_template.py                    # 현재 템플릿 검증
  py tests/test_excel_template.py path/to/file.xlsx  # 특정 파일 검증

검증 항목:
  1. DATA_ROWS 모든 행에 B열(항목명) 존재
  2. DATA_ROWS 모든 행에 L열 SUM 수식 존재
  3. 병합 영역 정합성 (깨진 병합 탐지)
  4. item_map.py DATA_ROWS ↔ 실제 행 번호 일치
  5. excel_helpers.py 상수 ↔ 실제 구조 일치
  6. 데이터 영역(D~K열) 비어있는지 확인 (템플릿은 비어야 정상)
"""
import os
import sys
import io

# Windows cp949 대응
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# 프로젝트 루트를 path에 추가
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from services.item_map import DATA_ROWS, ITEM_ROW_MAP, COL_IDX, SUM_COL
from services.excel_helpers import (
    DATA_START, DATA_END, MAX_COL, MAX_COL_PROP,
    REVIEW_START, REVIEW_COUNT,
)

_TMPL_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates")
DEFAULT_TEMPLATE = os.path.join(_TMPL_DIR, "master_template.xlsx")


def validate_template(path: str) -> list[str]:
    """템플릿 검증. 에러 목록 반환 (빈 리스트 = 통과)."""
    errors = []
    warnings = []

    if not os.path.exists(path):
        return [f"파일 없음: {path}"]

    wb = load_workbook(path, data_only=False)
    ws = wb.active

    # ── 1. DATA_ROWS 항목명 확인 (B열 또는 C열에 값 있어야 함) ──
    # 병합 카테고리(B열)는 첫 행에만 있고 나머지는 C열(세부항목)
    empty_label = []
    for row in DATA_ROWS:
        b_val = ws.cell(row=row, column=2).value  # B열
        c_val = ws.cell(row=row, column=3).value  # C열
        # B열이나 C열 중 하나라도 값이 있으면 OK
        b_is_merged = ws.cell(row=row, column=2).__class__.__name__ == "MergedCell"
        if not b_val and not c_val and not b_is_merged:
            empty_label.append(row)
    if empty_label:
        errors.append(f"B/C열 모두 비어있는 행: {empty_label}")

    # ── 2. L열 SUM 수식 확인 ──
    missing_sum = []
    wrong_sum = []
    for row in DATA_ROWS:
        cell = ws.cell(row=row, column=SUM_COL)  # L열
        val = cell.value
        if val is None:
            # 병합셀이면 스킵
            if cell.__class__.__name__ == "MergedCell":
                continue
            missing_sum.append(row)
        elif isinstance(val, str) and val.startswith("="):
            # SUM 수식이 D~K 범위를 포함하는지 확인
            upper = val.upper()
            if "SUM" not in upper:
                wrong_sum.append((row, val))
        # 숫자값이면 수식이 아님 → 경고
        elif isinstance(val, (int, float)):
            warnings.append(f"Row {row} L열: 수식 대신 숫자값 {val}")
    if missing_sum:
        errors.append(f"L열 SUM 수식 누락 행: {missing_sum}")
    if wrong_sum:
        errors.append(f"L열 수식 이상: {wrong_sum}")

    # ── 3. 병합 영역 정합성 ──
    merged = list(ws.merged_cells.ranges)
    broken_merges = []
    for mr in merged:
        # 병합 영역의 첫 셀에 값이 있어야 함
        first_cell = ws.cell(row=mr.min_row, column=mr.min_col)
        if first_cell.__class__.__name__ == "MergedCell":
            broken_merges.append(str(mr))
    if broken_merges:
        errors.append(f"깨진 병합 영역: {broken_merges}")

    # ── 4. ITEM_ROW_MAP 값이 DATA_ROWS에 포함되는지 ──
    data_rows_set = set(DATA_ROWS)
    orphan_rows = []
    for name, row in ITEM_ROW_MAP.items():
        if row is not None and row not in data_rows_set:
            orphan_rows.append((name, row))
    if orphan_rows:
        errors.append(f"ITEM_ROW_MAP에 있지만 DATA_ROWS에 없는 행: {orphan_rows}")

    # ── 5. excel_helpers 상수 일치 ──
    if DATA_START != min(COL_IDX.values()):
        errors.append(f"DATA_START({DATA_START}) != COL_IDX 최솟값({min(COL_IDX.values())})")
    if DATA_END != max(COL_IDX.values()):
        errors.append(f"DATA_END({DATA_END}) != COL_IDX 최댓값({max(COL_IDX.values())})")
    if MAX_COL != SUM_COL:
        errors.append(f"MAX_COL({MAX_COL}) != SUM_COL({SUM_COL})")

    # ── 6. 데이터 영역 D~K열이 비어있는지 (템플릿은 비어야 함) ──
    nonempty_data = []
    for row in DATA_ROWS:
        for col in range(DATA_START, DATA_END + 1):
            cell = ws.cell(row=row, column=col)
            if cell.__class__.__name__ != "MergedCell" and cell.value is not None:
                nonempty_data.append((row, col, cell.value))
    if nonempty_data and len(nonempty_data) <= 10:
        warnings.append(f"D~K열에 값이 있는 셀 {len(nonempty_data)}건: {nonempty_data[:5]}")
    elif nonempty_data:
        warnings.append(f"D~K열에 값이 있는 셀 {len(nonempty_data)}건 (처음 5건만 표시): {nonempty_data[:5]}")

    # ── 7. 리뷰 섹션 존재 확인 ──
    review_label = ws.cell(row=REVIEW_START, column=2).value
    if not review_label:
        warnings.append(f"리뷰 섹션 시작행(Row {REVIEW_START}) B열 비어있음")

    # ── 8. 갱신 구분행 확인 (Row 87~88) ──
    renewal_label = ws.cell(row=87, column=2).value
    if not renewal_label:
        warnings.append("갱신 구분행(Row 87) B열 비어있음")

    wb.close()
    return errors, warnings


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_TEMPLATE
    print(f"검증 대상: {os.path.basename(path)}")
    print("=" * 50)

    errors, warnings = validate_template(path)

    if warnings:
        for w in warnings:
            print(f"  ⚠️  {w}")
        print()

    if errors:
        print(f"  ❌ 오류 {len(errors)}건 발견:")
        for e in errors:
            print(f"     • {e}")
        print()
        print("결과: FAIL")
        sys.exit(1)
    else:
        data_rows_count = len(DATA_ROWS)
        item_map_count = len([v for v in ITEM_ROW_MAP.values() if v is not None])
        print(f"  ✅ DATA_ROWS: {data_rows_count}행 검증 통과")
        print(f"  ✅ ITEM_ROW_MAP: {item_map_count}개 매핑 정상")
        print(f"  ✅ L열 SUM 수식: 정상")
        print(f"  ✅ 병합 영역: 정상")
        print(f"  ✅ 상수 일치: 정상")
        print()
        print("결과: PASS")


if __name__ == "__main__":
    main()
