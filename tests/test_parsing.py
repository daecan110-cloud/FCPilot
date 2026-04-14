"""보장분석 PDF 파싱 회귀 테스트

사용법:
  python tests/test_parsing.py
  python tests/test_parsing.py --verbose

5개 PDF 기준 파싱 결과 검증:
  - 고객명/성별/나이
  - 계약 수, 보험사명
  - 보장항목 0값 없음
  - 주요 항목 기대값 일치
  - 엑셀 생성 + Row 검증
"""
import sys
import os
import io

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from services.pdf_extractor import extract_from_pdf
from services.excel_generator import generate_analysis_excel
from openpyxl import load_workbook
from services.item_map import DATA_ROWS

# ── PDF 경로 ──
_DL = os.path.expanduser("~/Downloads")
PDF_FILES = {
    "김세종": os.path.join(_DL, "김세종.pdf"),
    "전성언": os.path.join(_DL, "전성언.pdf"),
    "김은주": os.path.join(_DL, "김은주.pdf"),
    "김규연": os.path.join(_DL, "김규연.pdf"),
    "jin yexiang": os.path.join(_DL, "jin yexiang.pdf"),
}

# ── 기대값 ──
EXPECTED = {
    "김세종": {
        "고객명": "김세종", "성별": "남", "나이": 64,
        "계약수": 4,
        "보험사": ["DB생명보험", "메리츠화재보험", "메리츠화재보험", "DB손해보험"],
        "첫계약_월보": 98000,
        # 핵심 검증
        "재해사망_DB손해": ("DB손해보험", 12, 10000),  # (보험사, row, 값)
        "재해사망_메리츠운전자_없음": ("메리츠화재보험", "운전자보험 M-Drive", 12, 0),  # 메리츠 운전자에 재해사망 없어야
        "치아보철_없음": True,   # 어떤 계약에도 Row69 없어야
        "중증치매_없음": True,   # 어떤 계약에도 Row66 없어야
    },
    "전성언": {
        "고객명": "전성언", "성별": "남", "나이": 26,
        "계약수": 7,
        "보험사": ["흥국생명보험", "메리츠화재보험", "흥국화재해상보험",
                  "DB손해보험", "메리츠화재보험", "우정사업본부", "KB손해보험"],
        "첫계약_월보": 20537,
    },
    "김은주": {
        "고객명": "김은주", "성별": "여", "나이": 61,
        "계약수": 6,
        "첫계약_월보": 10000000,
        "치아보철_없음": True,
        "중증치매_없음": True,
    },
    "김규연": {
        "고객명": "김규연", "성별": "여", "나이": 32,
        "계약수": 5,
        "첫계약_월보": 10000000,
        "치아보철_없음": True,
    },
    "jin yexiang": {
        "고객명": "YEXIANG", "성별": "남", "나이": 32,
        "계약수": 14,   # 14개 추출 (엑셀에는 7개만)
        "첫계약_월보": 63808,
    },
}


def _load_pdf(name):
    path = PDF_FILES[name]
    if not os.path.exists(path):
        return None
    with open(path, "rb") as f:
        return extract_from_pdf(io.BytesIO(f.read()))


def test_basic_info(name, data, exp):
    """고객명/성별/나이 검증"""
    errors = []
    for key in ("고객명", "성별", "나이"):
        if key in exp and data.get(key) != exp[key]:
            errors.append(f"{key}: 기대={exp[key]}, 실제={data.get(key)}")
    return errors


def test_contracts(name, data, exp):
    """계약 수, 보험사명 검증"""
    errors = []
    contracts = data["_all_contracts"]
    if "계약수" in exp and len(contracts) != exp["계약수"]:
        errors.append(f"계약수: 기대={exp['계약수']}, 실제={len(contracts)}")
    if "보험사" in exp:
        actual = [c["보험사"] for c in contracts[:len(exp["보험사"])]]
        if actual != exp["보험사"]:
            errors.append(f"보험사 불일치")
    if "첫계약_월보" in exp:
        actual_prem = contracts[0]["월보험료"] if contracts else 0
        if actual_prem != exp["첫계약_월보"]:
            errors.append(f"첫계약 월보: 기대={exp['첫계약_월보']}, 실제={actual_prem}")
    return errors


def test_no_zero_values(name, data):
    """coverage_raw에 0값이 없는지 검증"""
    errors = []
    for ci, items in data["_coverage_raw"].items():
        for row, val in items.items():
            if val == 0:
                ct = data["_all_contracts"][ci]
                errors.append(f"0값: 계약{ci} {ct['보험사'][:10]} Row{row}")
    return errors


def test_key_items(name, data, exp):
    """핵심 보장항목 검증 (오매핑 방지)"""
    errors = []
    cov = data["_coverage_raw"]
    contracts = data["_all_contracts"]

    # 특정 계약의 특정 행 값 검증
    if "재해사망_DB손해" in exp:
        insurer, row, expected_val = exp["재해사망_DB손해"]
        for ci, items in cov.items():
            ct = contracts[ci]
            if insurer in ct["보험사"]:
                actual = items.get(str(row), 0)
                if actual != expected_val:
                    errors.append(f"재해사망 {insurer}: 기대={expected_val}, 실제={actual}")
                break

    # 메리츠 운전자보험에 재해사망 없어야 (교통상해사망 ≠ 재해사망)
    if "재해사망_메리츠운전자_없음" in exp:
        _, keyword, row, _ = exp["재해사망_메리츠운전자_없음"]
        for ci, items in cov.items():
            ct = contracts[ci]
            if keyword in ct["상품명"].replace("\n", "") and str(row) in items and items[str(row)] > 0:
                errors.append(f"메리츠 운전자보험에 재해사망 오매핑: {items[str(row)]}")

    # 치아보철 없어야
    if exp.get("치아보철_없음"):
        for ci, items in cov.items():
            if "69" in items and items["69"] > 0:
                ct = contracts[ci]
                errors.append(f"치아보철 오매핑: 계약{ci} {ct['보험사'][:10]}={items['69']}")

    # 중증치매 없어야
    if exp.get("중증치매_없음"):
        for ci, items in cov.items():
            if "66" in items and items["66"] > 0:
                ct = contracts[ci]
                errors.append(f"중증치매 오매핑: 계약{ci} {ct['보험사'][:10]}={items['66']}")

    return errors


def test_excel_generation(name, data):
    """엑셀 생성 + K열 SUM + 리뷰 검증"""
    errors = []
    try:
        result = generate_analysis_excel(data)
        fname, fbytes = result[0]
        wb = load_workbook(io.BytesIO(fbytes))
        ws = wb.active

        # 제목
        title = ws.cell(row=1, column=1).value or ""
        if not title or "보장 분석표" not in title:
            errors.append(f"제목 이상: {title[:30]}")

        # K열 SUM 수식
        for r in DATA_ROWS[:5]:
            v = ws.cell(row=r, column=11).value
            if not v or "=SUM" not in str(v):
                errors.append(f"K{r} SUM 누락")
                break

        # Row 5 보장기간 (2줄)
        r5 = ws.cell(row=5, column=4).value or ""
        if "\n" not in str(r5) and r5:
            errors.append(f"Row5 2줄 아님: {repr(r5)}")

        # Row 6 남은개월 (2줄)
        r6 = ws.cell(row=6, column=4).value or ""
        if "남은" not in str(r6) and r6:
            errors.append(f"Row6 남은개월 없음: {repr(r6)}")

        # Row 5/6 높이
        if ws.row_dimensions[5].height and ws.row_dimensions[5].height < 25:
            errors.append(f"Row5 높이 부족: {ws.row_dimensions[5].height}")
        if ws.row_dimensions[6].height and ws.row_dimensions[6].height < 25:
            errors.append(f"Row6 높이 부족: {ws.row_dimensions[6].height}")

        # 갱신형 (Row 80)
        contracts = data["_all_contracts"][:7]
        for i in range(len(contracts)):
            v = ws.cell(row=80, column=4 + i).value
            if not v:
                errors.append(f"Row80 갱신형 누락: col{4+i}")
                break

        # 리뷰 (Row 85~)
        for i in range(len(contracts)):
            v = ws.cell(row=85 + i, column=1).value
            if not v:
                errors.append(f"리뷰 Row{85+i} 누락")
                break

        # 기납입 보험료 — 납입완료 보험도 값 있어야
        for i, ct in enumerate(contracts):
            if ct.get("_paid", 0) > 0:
                v = ws.cell(row=75, column=4 + i).value
                if not v:
                    errors.append(f"기납입 누락: [{i}] {ct['보험사'][:10]}")
                    break

    except Exception as e:
        errors.append(f"엑셀 생성 실패: {e}")

    return errors


def run_all(verbose=False):
    total_pass = 0
    total_fail = 0
    total_skip = 0

    for name, exp in EXPECTED.items():
        print(f"\n{'='*50}")
        print(f"  {name}")
        print(f"{'='*50}")

        data = _load_pdf(name)
        if data is None:
            print(f"  SKIP — PDF 없음: {PDF_FILES[name]}")
            total_skip += 1
            continue

        all_errors = []

        # 테스트 실행
        tests = [
            ("기본정보", test_basic_info(name, data, exp)),
            ("계약", test_contracts(name, data, exp)),
            ("0값 없음", test_no_zero_values(name, data)),
            ("핵심항목", test_key_items(name, data, exp)),
            ("엑셀생성", test_excel_generation(name, data)),
        ]

        for test_name, errors in tests:
            if errors:
                all_errors.extend(errors)
                status = f"FAIL ({len(errors)})"
            else:
                status = "PASS"
            if verbose or errors:
                print(f"  [{status:8}] {test_name}")
                for e in errors:
                    print(f"             {e}")

        if all_errors:
            total_fail += 1
            print(f"  → FAIL: {len(all_errors)}개 오류")
        else:
            total_pass += 1
            if verbose:
                print(f"  → PASS")

    print(f"\n{'='*50}")
    print(f"  결과: PASS={total_pass} FAIL={total_fail} SKIP={total_skip}")
    print(f"{'='*50}")
    return total_fail == 0


if __name__ == "__main__":
    verbose = "--verbose" in sys.argv or "-v" in sys.argv
    ok = run_all(verbose=verbose)
    sys.exit(0 if ok else 1)
