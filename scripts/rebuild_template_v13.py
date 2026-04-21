"""v13 템플릿: v12 백업 기반으로 K열(합계)→L열 이동, K를 8번째 데이터 열로 전환"""
import os
import shutil
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(_DIR, "templates", "master_template_v12_backup.xlsx")
DST = os.path.join(_DIR, "templates", "master_template_v13.xlsx")

FONT_NAME = "KoPubWorld돋움체 Bold"
_THICK = Side(style="thick")

# 데이터 행 (v12와 동일)
DATA_ROWS = (
    list(range(9, 17)) + list(range(17, 35)) + list(range(35, 39))
    + list(range(39, 49)) + list(range(49, 54)) + list(range(54, 58))
    + list(range(58, 60)) + list(range(60, 71)) + list(range(71, 73))
    + [73] + list(range(74, 76)) + list(range(76, 78))
    + list(range(78, 82))
)

# 주요 섹션 경계 행 (이 행의 top에 medium 선)
MAJOR_DIVIDERS = [17, 35, 39, 54, 58, 71, 78, 82]


def build():
    shutil.copy(SRC, DST)
    wb = load_workbook(DST)
    ws = wb.active

    # ── 1. K열(col 11) 합계 서식+수식을 L열(col 12)로 복사 ──
    for r in range(1, 99):
        src = ws.cell(r, 11)  # K
        if src.__class__.__name__ == "MergedCell":
            continue
        dst = ws.cell(r, 12)  # L
        dst.value = src.value
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)

    # ── 2. K열을 데이터 열로 전환 (J열 서식 복사) ──
    for r in range(1, 99):
        src_j = ws.cell(r, 10)  # J
        if src_j.__class__.__name__ == "MergedCell":
            continue
        dst_k = ws.cell(r, 11)  # K
        if dst_k.__class__.__name__ == "MergedCell":
            continue
        dst_k.value = None  # 데이터 비우기
        dst_k.font = copy(src_j.font)
        dst_k.fill = copy(src_j.fill)
        dst_k.border = copy(src_j.border)
        dst_k.number_format = src_j.number_format
        dst_k.alignment = copy(src_j.alignment)

    # ── 3. 기존 K2:K6 병합 해제 ──
    for mc in list(ws.merged_cells.ranges):
        mc_str = str(mc)
        if "K2" in mc_str and "K6" in mc_str:
            ws.unmerge_cells(mc_str)
            break

    # ── 4. L열 합계 병합+라벨 ──
    ws.merge_cells("L2:L6")
    ws.cell(2, 12).value = "합계"
    ws.cell(2, 12).font = Font(name=FONT_NAME, size=9, bold=True)
    ws.cell(2, 12).alignment = Alignment(horizontal="center", vertical="center")

    # ── 5. L열 수식 전부 =SUM(D:K)로 변경 ──
    for row_num in DATA_ROWS:
        ws.cell(row_num, 12).value = f"=SUM(D{row_num}:K{row_num})"
    ws.cell(7, 12).value = "=SUM(D7:K7)"

    # 보험료 행
    ws.cell(82, 12).value = "=SUM(D82:K82)"
    ws.cell(83, 12).value = "=SUM(D83:K83)"
    ws.cell(84, 12).value = "=SUM(D84:K84)"

    # 총납입 = 기납입 + 남은 (D~K 각 열)
    for c in range(4, 12):  # D~K
        cl = get_column_letter(c)
        ws.cell(84, c).value = f"={cl}82+{cl}83"

    # ── 6. L열 너비 설정 ──
    ws.column_dimensions["K"].width = 13.0  # 데이터 열
    ws.column_dimensions["L"].width = 17.88  # 합계 열 (기존 K 너비)

    # ── 7. 단위 표시 이동 ──
    ws.cell(1, 11).value = None  # 기존 K1 단위 제거
    ws.cell(1, 12).value = "단위(만원)"
    ws.cell(1, 12).font = Font(name=FONT_NAME, size=8, bold=True)

    # ── 8. Row 1 병합 확장 (A1:J1 → A1:K1) ──
    for mc in list(ws.merged_cells.ranges):
        mc_str = str(mc)
        if mc_str == "A1:J1":
            ws.unmerge_cells(mc_str)
            ws.merge_cells("A1:K1")
            break

    # ── 9. 갱신형 구분 (Row 86~88) 병합 확장 ──
    for mc in list(ws.merged_cells.ranges):
        mc_str = str(mc)
        if mc_str == "A86:K86":
            ws.unmerge_cells(mc_str)
            ws.merge_cells("A86:L86")
            break

    # ── 10. 리뷰 섹션 (Row 90~98) 병합 확장 ──
    merge_updates = {
        "A90:K90": "A90:L90",
        "E91:H91": "E91:I91",
        "I91:K91": "J91:L91",
    }
    for old, new in merge_updates.items():
        for mc in list(ws.merged_cells.ranges):
            if str(mc) == old:
                ws.unmerge_cells(old)
                ws.merge_cells(new)
                break

    # 리뷰 행 병합 확장 (Row 92~99, 8상품 대응) + J:L fill 복사
    for r in range(92, 100):
        # E:H → E:I
        for mc in list(ws.merged_cells.ranges):
            mc_str = str(mc)
            if mc_str == f"E{r}:H{r}":
                ws.unmerge_cells(mc_str)
                ws.merge_cells(f"E{r}:I{r}")
                break
        # I:K → J:L  (I셀 fill/border를 J셀로 복사)
        old_anchor = ws.cell(r, 9)  # I (기존 anchor)
        old_fill = copy(old_anchor.fill)
        old_border = copy(old_anchor.border)
        old_font = copy(old_anchor.font)
        old_align = copy(old_anchor.alignment)
        for mc in list(ws.merged_cells.ranges):
            mc_str = str(mc)
            if mc_str == f"I{r}:K{r}":
                ws.unmerge_cells(mc_str)
                ws.merge_cells(f"J{r}:L{r}")
                break
        # J셀에 서식 복사
        new_anchor = ws.cell(r, 10)  # J (새 anchor)
        if new_anchor.__class__.__name__ != "MergedCell":
            new_anchor.fill = old_fill
            new_anchor.border = old_border
            new_anchor.font = old_font
            new_anchor.alignment = old_align

    # 리뷰 특이사항 열 이동
    ws.cell(91, 10).value = "특이사항 (면책기간, 보장범위 등)"

    # ── 11. 주요 섹션 경계에 medium 선 추가 ──
    for div_row in MAJOR_DIVIDERS:
        for c in range(2, 13):  # B~L
            cell = ws.cell(div_row, c)
            if cell.__class__.__name__ == "MergedCell":
                continue
            old = cell.border
            cell.border = Border(
                left=old.left, right=old.right,
                top=_THICK, bottom=old.bottom,
            )

    # ── 12. K열 헤더 행 서식 (Row 2~7) — J열 복사 ──
    for r in range(2, 8):
        src = ws.cell(r, 10)  # J
        if src.__class__.__name__ == "MergedCell":
            continue
        dst = ws.cell(r, 11)  # K
        if dst.__class__.__name__ == "MergedCell":
            continue
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)

    wb.save(DST)
    print("OK - v13 template built from v12 backup")


if __name__ == "__main__":
    build()
