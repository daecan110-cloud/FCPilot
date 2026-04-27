"""엑셀 생성 공통 헬퍼 — safe_val, 서식 복사, 셀 초기화, 병합"""
from copy import copy

from openpyxl.styles import Font, Alignment

_FONT_NAME = "KoPubWorld돋움체 Bold"

# v13 양식 공통 상수 — excel_generator / excel_review 공유
DATA_START = 4
DATA_END = 11
MAX_COL = 12
MAX_COL_PROP = 14
REVIEW_START = 92
REVIEW_COUNT = 7


def safe_val(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if cell.__class__.__name__ != "MergedCell":
        cell.value = value


def safe_merge(ws, range_str: str):
    """병합 생성. 겹치는 기존 병합이 있으면 해제 후 재생성."""
    from openpyxl.utils.cell import range_boundaries
    existing = {str(m) for m in ws.merged_cells.ranges}
    if range_str in existing:
        return
    # 겹치는 기존 병합 해제
    mc, mr, xc, xr = range_boundaries(range_str)
    to_remove = []
    for m in ws.merged_cells.ranges:
        if m.min_row <= xr and m.max_row >= mr and m.min_col <= xc and m.max_col >= mc:
            to_remove.append(str(m))
    for r in to_remove:
        try:
            ws.unmerge_cells(r)
        except Exception:
            pass
    try:
        ws.merge_cells(range_str)
    except Exception:
        pass


def copy_row_style(ws, src_row: int, dst_row: int, cols):
    """src_row의 셀 서식을 dst_row로 복사."""
    for c in cols:
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        if dst.__class__.__name__ == "MergedCell":
            continue
        if src.__class__.__name__ == "MergedCell":
            continue
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.font = copy(src.font)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def clear_values(ws, ranges: list):
    """지정된 (r_start, c_start, r_end, c_end) 범위의 셀 값을 None으로 초기화."""
    for r_s, c_s, r_e, c_e in ranges:
        for r in range(r_s, r_e + 1):
            for c in range(c_s, c_e + 1):
                cell = ws.cell(row=r, column=c)
                if cell.__class__.__name__ != "MergedCell":
                    cell.value = None


def classify_product_type(contract: dict) -> str:
    """상품명에서 상품 성격 분류 (보험사명은 제외)"""
    name = contract.get("상품명", "")

    if any(k in name for k in ["실손", "실비", "의료비"]):
        return "실손보험"
    if any(k in name for k in ["운전자", "드라이브", "drive", "Drive"]):
        return "운전자보험"
    if any(k in name for k in ["치아", "치과", "덴탈"]):
        return "치아보험"
    if any(k in name for k in ["종신"]):
        return "종신보험"
    if any(k in name for k in ["간병", "간호", "LTC"]):
        return "간병보험"
    if any(k in name for k in ["어린이", "자녀", "아이사랑"]):
        return "어린이보험"
    if any(k in name for k in ["저축", "연금", "변액"]):
        return "저축보험"
    if any(k in name for k in ["화재", "주택"]) and not any(
        k in name for k in ["건강", "종합", "케어", "플러스", "보장", "암", "치아",
                            "운전자", "실손", "의료비", "종신", "간병"]
    ):
        return "화재보험"
    if any(k in name for k in ["암", "癌"]) and not any(
        k in name for k in ["건강", "종합", "케어", "플러스", "보장"]
    ):
        return "암보험"
    if any(k in name for k in ["상해보험"]) and not any(
        k in name for k in ["건강", "종합"]
    ):
        return "상해보험"
    return "건강보험"


def short_name(contract: dict) -> str:
    """보험사·상품명 약칭"""
    name = contract.get("상품명", "")
    company = contract.get("보험사", "")
    for prefix in ["무배당 ", "(무배당)", "무배당", "無", "삼성 "]:
        name = name.replace(prefix, "")
    name = name.replace("()", "").strip()
    if "\n" in name:
        name = name.split("\n")[0]
    for full, short in [
        ("삼성생명보험", "삼성생명"), ("한화생명보험", "한화생명"),
        ("새마을금고중앙회", "새마을금고"), ("현대해상화재보험", "현대해상"),
    ]:
        company = company.replace(full, short)
    return f"{company}\n{name}"


def classify_renewal(ct: dict) -> tuple[str, str]:
    """계약 갱신 구분 + 보험료 변화 예고 판별. (renewal_text, notice_text) 반환."""
    name = ct.get("상품명", "")
    company = ct.get("보험사", "")
    prem = ct.get("월보험료", 0)
    total_m = ct.get("_총납입개월", 0)
    paid_m = ct.get("_납입개월", 0)

    is_renewal = "갱신" in name
    is_short = total_m and total_m <= 12
    is_sonhae = any(k in company for k in ["화재", "손해", "해상"])
    is_comprehensive = any(k in name for k in [
        "건강", "종합", "케어", "플러스", "훼밀리", "간편",
        "The", "NEW", "희망", "자녀", "Good",
    ])

    if is_renewal:
        renewal = "갱신형 ⚠️"
    elif is_short:
        renewal = "단기계약\n갱신없음"
    elif is_sonhae and is_comprehensive:
        renewal = "부분 갱신형 ⚠️"
    else:
        renewal = "비갱신형 ✅"

    if prem == 0:
        notice = "납입완료"
    elif is_renewal:
        notice = "갱신 시\n보험료 변동 예상 ⚠️"
    elif is_short:
        notice = "만기 소멸 예정"
    elif is_sonhae and is_comprehensive:
        notice = "특약 갱신 시\n일부 변동 가능 ⚠️"
    else:
        remain = total_m - paid_m if total_m else 0
        if remain <= 0:
            notice = "납입완료 예정"
        elif remain <= 24:
            notice = f"약 {remain}개월 후\n납입완료 예정"
        else:
            notice = "변동 없음"

    return renewal, notice
