"""보장 비교표 생성 — 복수 제안서 Before/After 비교"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from services.item_map import DATA_ROWS, ITEM_ROW_MAP
from services.proposal_parser import map_riders_to_rows


# ── 색상/스타일 ──
NAVY = "1B3A5C"
BLUE = "2E75B6"
WHITE = "FFFFFF"
VERY_LIGHT = "EBF5FB"
RED_BG = "F5B7B1"
YELLOW_BG = "F9E79F"
GREEN_BG = "ABEBC6"
GRAY_BG = "D5D8DC"
RED_FT = "C0392B"
GREEN_FT = "1E8449"
ORANGE_FT = "D68910"
GRAY_FT = "7F8C8D"
CAT_BG = "2C3E50"

thin = Side(style="thin", color="BDC3C7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 진단표 45개 항목 기준
_DIAG_ITEMS = [
    (9, "일반사망"), (10, "질병사망"), (11, "암사망"), (12, "재해사망"),
    (13, "재해장해(100%)"), (14, "재해장해(3%)"),
    (15, "질병장해(100%)"), (16, "질병장해(3%)"),
    (17, "일반암진단"), (18, "남녀특정암진단"),
    (19, "소액암진단"), (20, "고액암진단"),
    (21, "항암약물치료비"), (22, "표적항암약물허가치료비"),
    (23, "카티항암약물허가치료비"), (24, "항암방사선치료비"),
    (25, "항암세기조절방사선치료비"), (26, "항암양성자치료비"),
    (27, "항암중입자치료비"), (28, "암수술"),
    (29, "암로봇수술"),
    (30, "상급병원암주요치료(수술)"), (31, "상급병원암주요치료(약물)"),
    (32, "상급병원암주요치료(방사선)"), (33, "상급병원암주요치료(복합)"),
    (34, "비급여암주요치료비"),
    (35, "CI진단"),
    (36, "말기간질환진단"), (37, "말기폐질환진단"), (38, "말기신부전증진단"),
    (39, "뇌혈관질환진단"), (40, "뇌졸중진단"),
    (41, "뇌경색증진단"), (42, "뇌출혈진단"),
    (43, "허혈심장질환진단"), (44, "급성심근경색증진단"),
    (45, "뇌혈관수술"), (46, "허혈심장수술"),
    (47, "뇌출혈수술"), (48, "급성심근경색수술"),
    (49, "특정순환계-수술"), (50, "특정순환계-혈전용해"),
    (51, "특정순환계-혈전제거"), (52, "특정순환계-혈전복합"),
    (53, "특정순환계-중환자실"),
    (54, "질병수술(최소)"), (55, "재해수술"), (56, "질병수술(최대)"),
    (58, "골절진단비"), (59, "깁스치료비"),
    (60, "질병입원일당"), (61, "재해입원일당"), (62, "암입원일당"),
    (63, "종합병원입원(1인실)"), (64, "종합병원입원(2-3인실)"),
    (65, "상급종합병원입원(1인실)"), (66, "상급종합병원입원(2-3인실)"),
    (67, "중환자실입원비"),
    (68, "종합병원통원비"), (69, "상급종합병원통원비"),
    (70, "상급종합병원암통원"),
    (71, "경증치매진단"), (72, "중증치매(LTC)진단"), (73, "복합재가"),
    (74, "질병간병인사용일당"), (75, "상해간병인사용일당"),
    (76, "치아보철치료비"), (77, "치아보존치료비"),
    (78, "질병입원의료비"), (79, "질병통원의료비"),
    (80, "재해입원의료비"), (81, "재해통원의료비"),
]

_CATEGORIES = {
    9: "사망보장", 13: "후유장해", 17: "암보장",
    21: "암치료비·수술비", 30: "상급병원 암주요치료",
    35: "CI·말기질환·치매", 39: "뇌·심장 보장",
    49: "특정순환계질환", 54: "수술비", 58: "골절·깁스",
    60: "입원비", 68: "통원비", 71: "치매", 74: "간병비",
    76: "치아보장", 78: "실손의료비",
}


def generate_comparison_excel(
    data: dict,
    proposals: list[dict],
    selected_indices: list[int] | None = None,
) -> tuple[str, bytes]:
    """복수 제안서 비교표 엑셀 생성.

    Args:
        data: 보장분석 파싱 결과
        proposals: 제안서 파싱 결과 리스트 [{상품명, 보험료합계, 특약목록}, ...]
        selected_indices: 비교할 제안서 인덱스 (None이면 전부)

    Returns:
        (파일명, 엑셀 bytes)
    """
    if selected_indices is None:
        selected_indices = list(range(len(proposals)))

    sel_proposals = [proposals[i] for i in selected_indices if i < len(proposals)]
    n_props = len(sel_proposals)
    if n_props < 1:
        return ("", b"")

    # 현재 보장 합계
    current_sums = {}
    for ci, cov in data.get("_coverage_raw", {}).items():
        for k, v in cov.items():
            current_sums[k] = current_sums.get(k, 0) + v

    # 진단결과 (PDF에서 추출된 데이터)
    diag_data = data.get("_diag_results", {})

    # 각 제안서별 추가 금액 계산
    prop_adds = []
    for prop in sel_proposals:
        rider_map = map_riders_to_rows(prop.get("특약목록", []))
        adds = {}
        for row_num, amount in rider_map.items():
            adds[str(row_num)] = amount
        prop_adds.append(adds)

    # 제안서 이름 라벨
    labels = []
    for i, prop in enumerate(sel_proposals):
        name = prop.get("상품명", f"제안{i+1}")
        if len(name) > 15:
            name = name[:15] + "..."
        labels.append(name)

    # ── 엑셀 생성 ──
    wb = Workbook()
    ws = wb.active
    ws.title = "보장 비교표"

    # 열 구조: A(항목) B(진단) C(현재) [D(A안추가) E(A안후)] [F(B안추가) G(B안후)] H(필요) I(비고)
    base_cols = 3  # A항목 B진단 C현재
    prop_cols = n_props * 2  # 각 제안: 추가 + 변경후
    end_cols = 2  # 필요보장 + 비고
    total_cols = base_cols + prop_cols + end_cols

    # 열 너비
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 10
    for i in range(prop_cols):
        ws.column_dimensions[get_column_letter(4 + i)].width = 10
    ws.column_dimensions[get_column_letter(base_cols + prop_cols + 1)].width = 10
    ws.column_dimensions[get_column_letter(total_cols)].width = 20

    center = Alignment(horizontal="center", vertical="center")
    right_a = Alignment(horizontal="right", vertical="center")
    left_a = Alignment(horizontal="left", vertical="center", indent=1)
    data_font = Font(name="KoPubWorld돋움체 Bold", size=9)
    dash_font = Font(name="KoPubWorld돋움체 Bold", size=9, color="BDC3C7")
    plus_font = Font(name="KoPubWorld돋움체 Bold", size=9, color=GREEN_FT, bold=True)
    bold_font = Font(name="KoPubWorld돋움체 Bold", size=9, bold=True)

    diag_styles = {
        "부족": (PatternFill("solid", fgColor=RED_BG), Font(name="KoPubWorld돋움체 Bold", size=8, bold=True, color=RED_FT)),
        "미가입": (PatternFill("solid", fgColor=GRAY_BG), Font(name="KoPubWorld돋움체 Bold", size=8, bold=True, color=GRAY_FT)),
        "보통": (PatternFill("solid", fgColor=YELLOW_BG), Font(name="KoPubWorld돋움체 Bold", size=8, bold=True, color=ORANGE_FT)),
        "충분": (PatternFill("solid", fgColor=GREEN_BG), Font(name="KoPubWorld돋움체 Bold", size=8, bold=True, color=GREEN_FT)),
        "—": (PatternFill(), Font(name="KoPubWorld돋움체 Bold", size=8, color="95A5A6")),
    }

    def _all_border(r, ncols):
        for ci in range(1, ncols + 1):
            ws.cell(row=r, column=ci).border = border

    # ── Row 1: 타이틀 ──
    row = 1
    ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
    ws["A1"] = "보장 비교표"
    ws["A1"].font = Font(name="KoPubWorld돋움체 Bold", bold=True, size=16, color=NAVY)
    ws["A1"].alignment = center
    ws.row_dimensions[row].height = 38

    # ── Row 2: 고객 정보 ──
    row = 2
    score = data.get("보장점수", 0)
    info = f"{data.get('고객명', '')} 고객님 | {data.get('나이', '')}세 {data.get('성별', '')}성"
    if score:
        info += f" | 보장점수 {score}점"
    ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
    ws.cell(row=row, column=1, value=info).font = Font(
        name="맑은 고딕", size=10, color="566573")
    ws.cell(row=row, column=1).alignment = center
    ws.row_dimensions[row].height = 22

    # ── Row 3: 제안 상품 요약 ──
    row = 3
    prop_summary = " vs ".join(
        f"{labels[i]} (월 {sel_proposals[i].get('보험료합계', 0):,}원)"
        for i in range(n_props)
    )
    ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
    ws.cell(row=row, column=1, value=f"▶ {prop_summary}").font = Font(
        name="맑은 고딕", size=10, bold=True, color=BLUE)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=VERY_LIGHT)
    ws.cell(row=row, column=1).alignment = center
    ws.row_dimensions[row].height = 24

    # ── Row 4: 헤더 ──
    row = 4
    headers = ["보장항목", "진단", "현재 보장"]
    for i in range(n_props):
        letter = chr(65 + i)  # A, B, C...
        headers.append(f"{letter}안 추가")
        headers.append(f"{letter}안 변경후")
    headers.append("필요보장")
    headers.append("비고")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = Font(name="KoPubWorld돋움체 Bold", bold=True, color=WHITE, size=9)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[row].height = 22

    # ── 데이터 행 ──
    row = 5
    stripe = False
    last_cat = None

    for row_num, item_name in _DIAG_ITEMS:
        # 카테고리 헤더
        cat = None
        for cat_row, cat_name in _CATEGORIES.items():
            if row_num == cat_row:
                cat = cat_name
                break

        if cat and cat != last_cat:
            ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
            ws.cell(row=row, column=1, value=f"  {cat}").font = Font(
                name="맑은 고딕", bold=True, size=9, color=WHITE)
            for ci in range(1, total_cols + 1):
                ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=CAT_BG)
                ws.cell(row=row, column=ci).border = border
            ws.row_dimensions[row].height = 18
            last_cat = cat
            stripe = False
            row += 1

        cur = current_sums.get(str(row_num), 0)
        stripe = not stripe
        row_fill = PatternFill("solid", fgColor=VERY_LIGHT) if stripe else PatternFill()

        # 진단 결과: PDF 추출 데이터 우선, 없으면 미가입/— 판별
        diag = diag_data.get(str(row_num), "")
        if not diag:
            diag = "미가입" if cur == 0 else ""

        # A: 항목명
        ws.cell(row=row, column=1, value=item_name).font = data_font
        ws.cell(row=row, column=1).alignment = left_a
        ws.cell(row=row, column=1).fill = row_fill

        # B: 진단
        dfill, dfont = diag_styles.get(diag, (PatternFill(), data_font))
        ws.cell(row=row, column=2, value=diag).font = dfont
        ws.cell(row=row, column=2).fill = dfill if diag in diag_styles else row_fill
        ws.cell(row=row, column=2).alignment = center

        # C: 현재 보장
        c_cur = ws.cell(row=row, column=3)
        if cur > 0:
            c_cur.value = cur
            c_cur.number_format = "#,##0"
            c_cur.font = data_font
        else:
            c_cur.value = "—"
            c_cur.font = dash_font
        c_cur.alignment = right_a
        c_cur.fill = row_fill

        # D~: 각 제안서 (추가 + 변경후)
        for pi in range(n_props):
            add_val = prop_adds[pi].get(str(row_num), 0)
            after_val = cur + add_val
            col_add = base_cols + pi * 2 + 1
            col_after = base_cols + pi * 2 + 2

            c_add = ws.cell(row=row, column=col_add)
            if add_val > 0:
                c_add.value = add_val
                c_add.number_format = "+#,##0"
                c_add.font = plus_font
            else:
                c_add.value = "—"
                c_add.font = dash_font
            c_add.alignment = right_a
            c_add.fill = row_fill

            c_af = ws.cell(row=row, column=col_after)
            if after_val > 0:
                c_af.value = after_val
                c_af.number_format = "#,##0"
                c_af.font = bold_font
            else:
                c_af.value = "—"
                c_af.font = dash_font
            c_af.alignment = right_a
            c_af.fill = row_fill

        # 필요보장 (빈칸 — 진단표에서 가져올 수 있지만 현재 데이터에 없음)
        need_col = base_cols + prop_cols + 1
        ws.cell(row=row, column=need_col).fill = row_fill
        ws.cell(row=row, column=need_col).alignment = right_a

        # 비고
        note_col = total_cols
        ws.cell(row=row, column=note_col).fill = row_fill
        ws.cell(row=row, column=note_col).alignment = left_a

        _all_border(row, total_cols)
        ws.row_dimensions[row].height = 16
        row += 1

    # ── 하단 요약 ──
    row += 1
    ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
    ws.cell(row=row, column=1, value="  요약").font = Font(
        name="맑은 고딕", bold=True, size=10, color=WHITE)
    for ci in range(1, total_cols + 1):
        ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=row, column=ci).border = border
    ws.row_dimensions[row].height = 20
    row += 1

    # 월 보험료 비교
    cur_premium = sum(c.get("월보험료", 0) for c in data.get("_all_contracts", []))
    ws.cell(row=row, column=1, value="월 보험료").font = Font(
        name="맑은 고딕", size=9, bold=True, color=NAVY)
    ws.cell(row=row, column=1).alignment = left_a
    ws.cell(row=row, column=3, value=f"{cur_premium:,}").font = data_font
    ws.cell(row=row, column=3).alignment = right_a
    for pi in range(n_props):
        prop_prem = sel_proposals[pi].get("보험료합계", 0)
        col_add = base_cols + pi * 2 + 1
        col_after = base_cols + pi * 2 + 2
        ws.cell(row=row, column=col_add, value=f"+{prop_prem:,}").font = plus_font
        ws.cell(row=row, column=col_add).alignment = right_a
        ws.cell(row=row, column=col_after, value=f"{cur_premium + prop_prem:,}").font = bold_font
        ws.cell(row=row, column=col_after).alignment = right_a
    _all_border(row, total_cols)
    ws.row_dimensions[row].height = 20

    # ── 범례 ──
    row += 2
    ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
    ws.cell(row=row, column=1,
            value="(단위: 만원) | — 해당없음/변동없음").font = Font(
        name="맑은 고딕", size=7, color="95A5A6")
    ws.cell(row=row, column=1).alignment = center

    # ── 인쇄 설정 ──
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    buf = io.BytesIO()
    wb.save(buf)

    customer = data.get("고객명", "고객")
    filename = f"{customer}_보장비교표.xlsx"
    return (filename, buf.getvalue())
