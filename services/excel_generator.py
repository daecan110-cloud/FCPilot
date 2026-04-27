"""보장분석 엑셀 생성기 — v13 양식 8상품 (2026-04)"""
import io
import os
import shutil
import tempfile
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.pagebreak import Break

from services.item_map import COL_IDX, COL_LTRS, SUM_COL, PROP_COL, TOTAL_COL, DATA_ROWS
from services.excel_helpers import (
    safe_val, clear_values, _FONT_NAME, classify_product_type,
    DATA_START as _DATA_START, DATA_END as _DATA_END,
    MAX_COL as _MAX_COL, MAX_COL_PROP as _MAX_COL_PROP,
    REVIEW_START as _REVIEW_START, REVIEW_COUNT as _REVIEW_COUNT,
)
from services.excel_review import (
    fill_renewal, fill_renewal_all, fill_review, fill_review_all,
)

_TMPL_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates")
TEMPLATE = os.path.join(_TMPL_DIR, "master_template.xlsx")


def generate_analysis_excel(
    data: dict, proposal: dict | None = None,
    review_last: bool = False, **_kw,
) -> list[tuple[str, bytes]]:
    all_contracts = data.get("_all_contracts", data.get("계약", []))
    coverage_raw = data.get("_coverage_raw", {})
    customer = data.get("고객명", "고객")

    chunks = [all_contracts[i:i + 8] for i in range(0, len(all_contracts), 8)]
    total_pages = len(chunks)
    results: list[tuple[str, bytes]] = []

    for page, chunk in enumerate(chunks):
        sd = _make_slice(data, chunk, coverage_raw)
        prop = proposal if page == 0 else None
        is_last = (page == total_pages - 1)

        if review_last and total_pages > 1:
            review_contracts = all_contracts if is_last else []
        else:
            review_contracts = None

        b = _fill_workbook(
            sd, proposal=prop,
            review_contracts=review_contracts,
            all_coverage_raw=coverage_raw if review_contracts else None,
        )
        suffix = f"_{page + 1}" if total_pages > 1 else ""
        results.append((f"{customer}_보장분석표{suffix}.xlsx", b))

    return results


def _make_slice(base, contracts, coverage_raw):
    data = {
        "고객명": base["고객명"], "성별": base["성별"], "나이": base["나이"],
        "계약": [], "보장금액": {},
    }
    for new_i, c in enumerate(contracts):
        col = COL_LTRS[new_i]
        nc = {k: v for k, v in c.items() if not k.startswith("_")}
        nc["열"] = col
        nc["_납입기간"] = c.get("_납입기간", "")
        nc["_납입개월"] = c.get("_납입개월", 0)
        nc["_총납입개월"] = c.get("_총납입개월", 0)
        nc["_paid"] = c.get("_paid", 0)
        nc["_topay"] = c.get("_topay", 0)
        data["계약"].append(nc)
        if c["_idx"] in coverage_raw:
            data["보장금액"][col] = coverage_raw[c["_idx"]]
    return data


def _fill_workbook(slice_data, proposal=None, review_contracts=None, all_coverage_raw=None):
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    shutil.copy(TEMPLATE, tmp.name)
    wb = load_workbook(tmp.name)
    ws = wb.active

    for ns in wb._named_styles:
        if ns.name == "Normal":
            ns.font = Font(name=_FONT_NAME, size=10, bold=True)
            break

    has_proposal = proposal and proposal.get("특약목록")
    contracts = slice_data.get("계약", [])
    _clear_values(ws, has_proposal=has_proposal)
    _fill_header(ws, slice_data)
    _fill_coverage(ws, slice_data)
    if has_proposal:
        _fill_proposal(ws, proposal)
        _format_proposal_cols(ws)
    _fill_sums(ws, contracts, has_proposal=has_proposal, proposal=proposal)

    # 실제 상품 수 기반 열 숨기기 + 리뷰 병합 범위 계산
    n_contracts = len(contracts)
    visible_end = max(n_contracts, 4)  # 최소 4열 유지
    last_data_col = _DATA_START + visible_end - 1  # 마지막 보이는 데이터 열

    if review_contracts is None:
        fill_renewal(ws, contracts)
        fill_review(ws, contracts, slice_data.get("보장금액", {}),
                    last_data_col=last_data_col)
    elif len(review_contracts) > 0:
        fill_renewal_all(ws, review_contracts, last_data_col=last_data_col)
        fill_review_all(ws, review_contracts, all_coverage_raw or {},
                        last_data_col=last_data_col)

    _final_format(ws, has_proposal=has_proposal)
    _hide_unused_columns(ws, n_contracts, has_proposal=has_proposal)
    _clear_unused_review_rows(ws, n_contracts)

    buf = io.BytesIO()
    wb.save(buf)
    os.unlink(tmp.name)
    return buf.getvalue()


def _hide_unused_columns(ws, n_contracts: int, has_proposal: bool = False):
    """사용하지 않는 데이터 열 숨기기 + 열 너비 균등 분배"""
    from openpyxl.utils import get_column_letter
    visible = n_contracts
    # 열 너비: D~K 총 너비를 보이는 열 수로 균등 분배
    total_width = sum(
        ws.column_dimensions[get_column_letter(_DATA_START + i)].width
        for i in range(8)
    )
    even_width = total_width / max(visible, 4)
    for i in range(visible):
        col_letter = get_column_letter(_DATA_START + i)
        ws.column_dimensions[col_letter].width = even_width
    hidden_any = False
    for i in range(visible, 8):  # 미사용 열 숨기기
        col_idx = _DATA_START + i
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].hidden = True
        hidden_any = True
    if hidden_any:
        ws.column_dimensions["L"].width = 25


def _clear_unused_review_rows(ws, n_contracts: int):
    """리뷰 섹션에서 계약 수 초과 행의 서식(fill/border) 제거"""
    from openpyxl.styles import PatternFill, Border
    empty_fill = PatternFill(fill_type=None)
    empty_border = Border()
    # 템플릿에 fill이 있는 행 전체 커버 (Row 92~99 = 8행)
    for r in range(_REVIEW_START + n_contracts, _REVIEW_START + 8):
        for c in range(1, _MAX_COL_PROP + 1):
            cell = ws.cell(row=r, column=c)
            if cell.__class__.__name__ == "MergedCell":
                continue
            cell.value = None
            cell.fill = empty_fill
            cell.border = empty_border


def _clear_values(ws, has_proposal=False):
    max_c = _MAX_COL_PROP if has_proposal else _MAX_COL
    ranges = [
        (1, 1, 1, max_c),
        (3, _DATA_START, 7, _DATA_END),
        (9, _DATA_START, 81, _DATA_END),
        (9, SUM_COL, 81, SUM_COL),
        (82, _DATA_START, 84, max_c),
        (_REVIEW_START, 1, _REVIEW_START + _REVIEW_COUNT - 1, max_c),
    ]
    if has_proposal:
        ranges += [
            (2, PROP_COL, 7, PROP_COL),
            (9, PROP_COL, 81, PROP_COL),
            (2, TOTAL_COL, 7, TOTAL_COL),
            (9, TOTAL_COL, 81, TOTAL_COL),
        ]
    clear_values(ws, ranges)


def _fill_header(ws, slice_data):
    customer = slice_data.get("고객명", "고객")
    gender = slice_data.get("성별", "남")
    age = slice_data.get("나이", 0)
    gender_full = "남자" if gender == "남" else "여자"
    safe_val(ws, 1, 1, f"{customer}님 ({gender_full}, {age}세) · 보장 분석표")

    safe_val(ws, 4, 3, "상품성격")
    safe_val(ws, 5, 3, "가입시기 / 납입기간\n(보장기간)")
    for c in slice_data.get("계약", []):
        col = COL_IDX.get(c["열"], 4)
        safe_val(ws, 2, col, c.get("보험사", ""))
        safe_val(ws, 3, col, c.get("상품명", ""))
        safe_val(ws, 4, col, classify_product_type(c))
        가입시기 = c.get("가입시기", "")
        납입기간 = c.get("_납입기간", "")
        coverage_period = c.get("보장나이", "")
        parts = [p for p in [가입시기, 납입기간] if p]
        line1 = " / ".join(parts)
        if line1 and coverage_period:
            safe_val(ws, 5, col, f"{line1}\n({coverage_period})")
        elif line1:
            safe_val(ws, 5, col, line1)
        else:
            safe_val(ws, 5, col, coverage_period or None)
        paid_m = c.get("_납입개월", 0)
        total_m = c.get("_총납입개월", 0)
        if total_m:
            remain = total_m - paid_m
            safe_val(ws, 6, col, f"{paid_m}/{total_m}\n(남은 {remain}개월)")
        safe_val(ws, 7, col, c.get("월보험료", 0))


def _fill_coverage(ws, slice_data):
    for col_ltr, row_data in slice_data.get("보장금액", {}).items():
        col = COL_IDX.get(col_ltr)
        if not col:
            continue
        for row_str, amount in row_data.items():
            row_num = int(row_str)
            if row_num in DATA_ROWS:
                safe_val(ws, row_num, col, amount if amount else None)


def _fill_proposal(ws, proposal):
    """L열(col 12)에 신상품 제안 데이터 채우기"""
    from services.proposal_parser import map_riders_to_rows

    riders = proposal.get("특약목록", [])
    주계약 = riders[0] if riders else {}

    safe_val(ws, 2, PROP_COL, "신상품 제안")
    safe_val(ws, 3, PROP_COL, proposal.get("상품명", "제안 상품"))
    납입 = 주계약.get("납입기간", "")
    보장 = 주계약.get("보험기간", "")
    if 납입 and 보장:
        safe_val(ws, 5, PROP_COL, f"{납입}\n({보장})")
    elif 납입:
        safe_val(ws, 5, PROP_COL, 납입)
    else:
        safe_val(ws, 5, PROP_COL, 보장)
    total_prem = proposal.get("보험료합계", 0)
    if total_prem:
        safe_val(ws, 7, PROP_COL, total_prem)

    row_amounts = map_riders_to_rows(riders)
    for row_num, amount in row_amounts.items():
        if row_num in DATA_ROWS:
            safe_val(ws, row_num, PROP_COL, amount)


def _format_proposal_cols(ws):
    """L열(합계) 서식을 M, N열에 복사 + 셀 병합"""
    from openpyxl.utils import get_column_letter

    l_ltr = get_column_letter(PROP_COL)
    m_ltr = get_column_letter(TOTAL_COL)

    ws.column_dimensions[l_ltr].width = 18
    ws.column_dimensions[m_ltr].width = 18

    for r in range(1, 87):
        src = ws.cell(row=r, column=SUM_COL)
        if src.__class__.__name__ == "MergedCell":
            continue
        for col in (PROP_COL, TOTAL_COL):
            dst = ws.cell(row=r, column=col)
            if dst.__class__.__name__ == "MergedCell":
                continue
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)

    src2 = ws.cell(row=2, column=SUM_COL)
    _thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    _hdr_style = {
        "font": Font(name=_FONT_NAME, size=9, bold=True),
        "fill": copy(src2.fill),
        "border": _thin,
        "number_format": "#,##0",
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }
    for r in range(2, 7):
        for col in (PROP_COL, TOTAL_COL):
            cell = ws.cell(row=r, column=col)
            if cell.__class__.__name__ != "MergedCell":
                for attr, val in _hdr_style.items():
                    setattr(cell, attr, copy(val) if attr != "number_format" else val)

    ws.merge_cells(f"{l_ltr}3:{l_ltr}4")
    ws.merge_cells(f"{l_ltr}5:{l_ltr}6")
    ws.merge_cells(f"{m_ltr}2:{m_ltr}6")


def _fill_sums(ws, contracts, has_proposal=False, proposal=None):
    from openpyxl.utils import get_column_letter
    start_ltr = get_column_letter(_DATA_START)
    end_ltr = get_column_letter(_DATA_END)
    k_ltr = get_column_letter(SUM_COL)
    l_ltr = get_column_letter(PROP_COL)

    for row_num in DATA_ROWS:
        safe_val(ws, row_num, SUM_COL, f"=SUM({start_ltr}{row_num}:{end_ltr}{row_num})")
    safe_val(ws, 7, SUM_COL, f"=SUM({start_ltr}7:{end_ltr}7)")

    if has_proposal:
        safe_val(ws, 2, TOTAL_COL, "전체합계")
        for row_num in DATA_ROWS:
            safe_val(ws, row_num, TOTAL_COL, f"={k_ltr}{row_num}+{l_ltr}{row_num}")
        safe_val(ws, 7, TOTAL_COL, f"={k_ltr}7+{l_ltr}7")

    for ct in contracts:
        col = COL_IDX.get(ct["열"], 4)
        paid_amt = ct.get("_paid", 0)
        if not paid_amt:
            prem = ct.get("월보험료", 0)
            paid_m = ct.get("_납입개월", 0)
            paid_amt = int(prem * paid_m) if prem and paid_m else 0
        if paid_amt:
            safe_val(ws, 82, col, paid_amt)
    safe_val(ws, 82, SUM_COL, f"=SUM({start_ltr}82:{end_ltr}82)")

    for ct in contracts:
        col = COL_IDX.get(ct["열"], 4)
        topay_amt = ct.get("_topay", 0)
        if not topay_amt:
            prem = ct.get("월보험료", 0)
            total_m = ct.get("_총납입개월", 0)
            paid_m = ct.get("_납입개월", 0)
            remain = total_m - paid_m
            topay_amt = int(prem * remain) if prem and remain > 0 else 0
        if topay_amt:
            safe_val(ws, 83, col, topay_amt)
    safe_val(ws, 83, SUM_COL, f"=SUM({start_ltr}83:{end_ltr}83)")

    for c_ltr in COL_LTRS:
        col = COL_IDX[c_ltr]
        col_l = get_column_letter(col)
        safe_val(ws, 84, col, f"={col_l}82+{col_l}83")
    safe_val(ws, 84, SUM_COL, f"=SUM({start_ltr}84:{end_ltr}84)")

    if has_proposal:
        from services.proposal_parser import _parse_납입개월
        if proposal:
            riders = proposal.get("특약목록", [])
            주계약 = riders[0] if riders else {}
            납입개월 = _parse_납입개월(주계약.get("납입기간", ""))
            total_prem = proposal.get("보험료합계", 0)
            if 납입개월 and total_prem:
                topay = total_prem * 납입개월
                safe_val(ws, 83, PROP_COL, topay)
                safe_val(ws, 84, PROP_COL, topay)

        safe_val(ws, 82, TOTAL_COL, f"={k_ltr}82+{l_ltr}82")
        safe_val(ws, 83, TOTAL_COL, f"={k_ltr}83+{l_ltr}83")
        safe_val(ws, 84, TOTAL_COL, f"={k_ltr}84+{l_ltr}84")


def _final_format(ws, has_proposal=False):
    max_c = _MAX_COL_PROP if has_proposal else _MAX_COL
    for r in range(1, ws.max_row + 1):
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            if cell.__class__.__name__ == "MergedCell":
                continue
            old = cell.font
            cell.font = Font(
                name=_FONT_NAME,
                size=old.size if old.size else 9,
                bold=True,
                italic=old.italic if old.italic else False,
                color=old.color,
            )
        if r in (3, 5, 6):
            for c in range(_DATA_START, _DATA_END + 1):
                cell = ws.cell(row=r, column=c)
                if cell.__class__.__name__ != "MergedCell":
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True,
                    )
    if has_proposal:
        for r in (3, 5, 6):
            cell = ws.cell(row=r, column=PROP_COL)
            if cell.__class__.__name__ != "MergedCell":
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True,
                )
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 30
    ws.row_dimensions[8].height = 10
    ws.row_dimensions[85].height = 5
    for r in range(_REVIEW_START, _REVIEW_START + _REVIEW_COUNT):
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            if cell.__class__.__name__ != "MergedCell":
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True,
                )
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.2
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.scale = None
    ws.row_breaks.brk = []
    ws.row_breaks.append(Break(id=85))
