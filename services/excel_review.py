"""엑셀 보장분석 — 리뷰/갱신 섹션 (excel_generator에서 분리)"""
import math
import re

from copy import copy
from openpyxl.styles import Font, Alignment, Border, PatternFill

from services.item_map import COL_IDX, COL_LTRS, DATA_ROWS
from services.excel_helpers import (
    safe_val, safe_merge, copy_row_style, short_name, classify_renewal,
    _FONT_NAME,
)

_DATA_START = 4
_DATA_END = 11
_MAX_COL = 12
_MAX_COL_PROP = 14
_REVIEW_START = 92
_REVIEW_COUNT = 7


# ── 갱신 구분 ────────────────────────────────────────────

def fill_renewal(ws, contracts):
    """Row 87 갱신 구분, Row 88 보험료 변화 예고"""
    for ct in contracts:
        col = COL_IDX.get(ct["열"], 4)
        renewal, notice = classify_renewal(ct)
        safe_val(ws, 87, col, renewal)
        safe_val(ws, 88, col, notice)
    # L열 fill을 데이터 열과 통일
    _sync_renewal_L_fill(ws, [87, 88])


def fill_renewal_all(ws, all_contracts, last_data_col: int = _DATA_END):
    """전체 계약 갱신/보험료변화 (8개씩 그룹, 동적 행 확장)"""
    n_groups = math.ceil(len(all_contracts) / 8)
    check_end, note_start = _review_merge_cols(last_data_col)

    if n_groups > 1:
        extra_rows = (n_groups - 1) * 2
        # 병합 완전 해제 후 insert
        _force_unmerge_range(ws, 86, 100)
        ws.insert_rows(89, extra_rows)
        _force_unmerge_range(ws, 86, 100 + extra_rows)

        for i in range(extra_rows):
            r_new = 89 + i
            r_src = 87 + (i % 2)
            copy_row_style(ws, r_src, r_new, cols=range(1, _MAX_COL_PROP + 1))
            src_h = ws.row_dimensions[r_src].height if ws.row_dimensions.get(r_src) else None
            if src_h:
                ws.row_dimensions[r_new].height = src_h

        # 갱신 타이틀 재생성
        ws.merge_cells("A86:L86")
        safe_val(ws, 86, 1, "🔄  갱신형 / 비갱신형 구분")

        review_title_row = 90 + extra_rows
        review_hdr_row = review_title_row + 1
        _write_review_header(ws, review_title_row, review_hdr_row,
                             check_end, note_start)

    for g in range(n_groups):
        chunk = all_contracts[g * 8:(g + 1) * 8]
        renewal_row = 87 + g * 2
        notice_row = 88 + g * 2

        for i, ct in enumerate(chunk):
            col = COL_IDX.get(COL_LTRS[i], 4)
            renewal, notice = classify_renewal(ct)
            safe_val(ws, renewal_row, col, renewal)
            safe_val(ws, notice_row, col, notice)

        renewal_rows = []
        for row in (renewal_row, notice_row):
            renewal_rows.append(row)
            for c in range(_DATA_START, _DATA_END + 1):
                cell = ws.cell(row=row, column=c)
                if cell.__class__.__name__ != "MergedCell":
                    cell.font = Font(name=_FONT_NAME, size=8, bold=True)
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True,
                    )
            ws.row_dimensions[row].height = 26 if row % 2 == 0 else 20
        # L열 fill을 데이터 열과 통일
        _sync_renewal_L_fill(ws, renewal_rows)

    return (n_groups - 1) * 2


# ── 리뷰 ─────────────────────────────────────────────────

def fill_review(ws, contracts, coverage_map=None, last_data_col: int = _DATA_END):
    """슬라이스 계약 리뷰 (병합+서식 직접 적용)"""
    n = len(contracts)
    if n == 0:
        return
    # 교대 fill 패턴 읽기 (병합 해제 전에)
    fills, borders = _read_review_fills(ws)

    check_end, note_start = _review_merge_cols(last_data_col)

    # 기존 병합 해제 (헤더 + 데이터 영역)
    _force_unmerge_range(ws, _REVIEW_START - 2, _REVIEW_START + n + 2)
    # 헤더 병합 재조정 (템플릿 기본 헤더)
    _write_review_header(ws, _REVIEW_START - 2, _REVIEW_START - 1,
                         check_end, note_start)

    for i, c in enumerate(contracts):
        r = _REVIEW_START + i
        # fill 먼저 적용 (병합 전에 모든 셀에 배경색)
        _apply_review_style(ws, r, i, fills, borders)
        # 병합 생성 (동적 범위)
        _merge_review_row(ws, r, check_end, note_start)

        cov = (coverage_map or {}).get(c.get("열", ""), {})
        _write_review_row(ws, r, c, cov, note_col=note_start)
        ws.row_dimensions[r].height = 70


def fill_review_all(ws, all_contracts, all_coverage_raw, last_data_col: int = _DATA_END):
    """전체 계약 리뷰 (insert_rows 없이 직접 쓰기)"""
    n_groups = math.ceil(len(all_contracts) / 8)
    extra_renewal = (n_groups - 1) * 2 if n_groups > 1 else 0

    review_start = _REVIEW_START + extra_renewal
    n_contracts = len(all_contracts)
    check_end, note_start = _review_merge_cols(last_data_col)

    # 리뷰 영역 전체 병합 해제 (기존 병합 완전 제거)
    _force_unmerge_range(ws, review_start - 2, review_start + n_contracts + 5)

    # 타이틀 + 헤더
    title_row = review_start - 2
    header_row = review_start - 1
    _write_review_header(ws, title_row, header_row, check_end, note_start)

    # 교대 fill 패턴 읽기 (review_start 기준, insert_rows 반영)
    fills, borders = _read_review_fills(ws, review_start)

    # 각 계약 리뷰 행 직접 쓰기
    for i, c in enumerate(all_contracts):
        r = review_start + i
        # fill 먼저 적용 (병합 전에 모든 셀에 배경색)
        _apply_review_style(ws, r, i, fills, borders)
        # 병합 생성 (동적 범위)
        _merge_review_row(ws, r, check_end, note_start)

        # 데이터 쓰기
        cov_data = {}
        if c["_idx"] in all_coverage_raw:
            cov_data = {str(k): v for k, v in all_coverage_raw[c["_idx"]].items()}
        _write_review_row(ws, r, c, cov_data, note_col=note_start)
        ws.row_dimensions[r].height = 70


def _read_review_fills(ws, start_row: int = 92):
    """템플릿에서 리뷰 행 교대 fill/border 패턴을 읽어 캐시.
    start_row: 첫 번째 리뷰 데이터 행 (insert_rows 반영 후)"""
    fills = []
    borders = []
    for ref_row in (start_row, start_row + 1):
        cell = ws.cell(row=ref_row, column=1)
        if cell.__class__.__name__ != "MergedCell":
            fills.append(copy(cell.fill))
            borders.append(copy(cell.border))
        else:
            fills.append(None)
            borders.append(None)
    return fills, borders


def _apply_review_style(ws, r: int, idx: int, fills, borders):
    """리뷰 행 전체 서식 강제 적용 — font, alignment, fill, border (A~L).
    idx: 0-based 행 인덱스 (짝수=fills[0], 홀수=fills[1])"""
    _font = Font(name=_FONT_NAME, size=8, bold=True)
    _align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ref_fill = fills[idx % 2] if fills else None
    ref_border = borders[idx % 2] if borders else None

    # 모든 열 서식 적용 (A~L, col 1~12)
    for col_n in range(1, _MAX_COL + 1):
        cell = ws.cell(row=r, column=col_n)
        if cell.__class__.__name__ == "MergedCell":
            # ghost MergedCell 삭제 후 새 셀 생성
            if (r, col_n) in ws._cells:
                del ws._cells[(r, col_n)]
            cell = ws.cell(row=r, column=col_n)
            # 다시 MergedCell이면 스킵
            if cell.__class__.__name__ == "MergedCell":
                continue
        cell.font = _font
        cell.alignment = _align
        if ref_fill:
            cell.fill = copy(ref_fill)
        if ref_border:
            cell.border = copy(ref_border)


def _write_review_row(ws, r: int, c: dict, cov_data: dict, note_col: int = 10):
    """리뷰 행 1줄 쓰기 (fill_review / fill_review_all 공통)"""
    safe_val(ws, r, 1, short_name(c))
    period = c.get("보장나이", "")
    start = c.get("가입시기", "")
    safe_val(ws, r, 3, f"{start}\n({period})" if start else period)
    prem = c.get("월보험료", 0)
    safe_val(ws, r, 4, f"{prem:,.0f}원" if prem else "납입완료")
    safe_val(ws, r, 5, build_review(c, coverage_data=cov_data))


# ── 리뷰 텍스트 생성 ─────────────────────────────────────

def build_review(contract, coverage_data=None):
    name = contract.get("상품명", "")
    company = contract.get("보험사", "")
    combined = name + company
    period = contract.get("보장나이", "")
    premium = contract.get("월보험료", 0)
    checks = []

    if any(k in combined for k in ["단체", "단기"]):
        checks.append("단체/단기계약 — 퇴직·탈퇴 시 자동 소멸")
    elif any(k in combined for k in ["실손", "의료비"]):
        start_date = contract.get("가입시기", "")
        gen = _detect_silbi_gen(name, company, start_date)
        checks.append(f"실손의료비 {gen} (갱신형)")
    elif any(k in combined for k in ["종신"]) and any(k in combined for k in ["변액", "유니버셜"]):
        checks.append("변액유니버셜종신 — 수익률·적립금 확인 필요")
    elif any(k in combined for k in ["종신"]):
        checks.append("종신보험 — 비갱신형")
    elif any(k in combined for k in ["운전자", "자동차"]):
        checks.append("운전자보험 — 교통상해/벌금/변호사비 보장")
    elif any(k in combined for k in ["CI"]):
        checks.append("CI보험 — 중대질병 진단 시 보험금 지급")
    elif any(k in combined for k in ["치아"]):
        checks.append(f"치아보험 ({period})")
    elif any(k in combined for k in ["건강", "상해", "종합"]):
        checks.append(f"건강/상해 보장 ({period})")
    elif any(k in combined for k in ["암"]):
        checks.append(f"암보험 ({period})")
    elif any(k in combined for k in ["저축", "연금", "적립"]):
        checks.append(f"저축/연금 ({period})")
    else:
        checks.append(f"보장기간: {period}")

    if premium == 0:
        checks.append("납입완료")

    if coverage_data:
        highlights = _coverage_highlights(coverage_data)
        if highlights:
            checks.append(highlights)

    return " / ".join(checks)


def _detect_silbi_gen(name: str, company: str, start_date: str = "") -> str:
    """실손의료비 세대 판별"""
    combined = name + company

    if any(k in combined for k in ["착한", "비급여특약"]):
        return "4세대"
    if any(k in combined for k in ["4세대", "신실손"]):
        return "4세대"
    if any(k in combined for k in ["표준화", "3세대"]):
        return "3세대"
    if any(k in combined for k in ["단독", "2세대"]):
        return "2세대"
    if any(k in combined for k in ["1세대"]):
        return "1세대"
    if any(k in combined for k in ["5세대"]):
        return "5세대"

    m = re.search(r"(20[0-2]\d)", name)
    if m:
        year = int(m.group(1))
        if year >= 2026:
            return "5세대"
        if year >= 2017:
            return "4세대"

    if start_date:
        m = re.match(r"(\d{4})-?(\d{2})?", start_date)
        if m:
            y = int(m.group(1))
            mo = int(m.group(2) or "1")
            ym = y * 100 + mo
            if ym >= 202605:
                return "5세대"
            if ym >= 201704:
                return "4세대"
            if ym >= 201301:
                return "3세대"
            if ym >= 200910:
                return "2세대"
            return "1세대"

    return "(세대 확인 필요)"


def _coverage_highlights(cov: dict) -> str:
    """주요 보장 항목 요약"""
    labels = {
        "9": "사망", "17": "일반암", "44": "심근경색",
        "39": "뇌혈관", "78": "실비",
    }
    parts = []
    for row_str, label in labels.items():
        amt = cov.get(row_str, 0)
        if amt:
            parts.append(f"{label} {amt:,}")
    return " / ".join(parts) if parts else ""


def _sync_renewal_L_fill(ws, rows: list):
    """갱신 행의 L열(col 12) fill/font/alignment을 D열과 통일."""
    for row in rows:
        ref = ws.cell(row=row, column=4)  # D열
        if ref.__class__.__name__ == "MergedCell":
            continue
        target = ws.cell(row=row, column=12)  # L열
        if target.__class__.__name__ == "MergedCell":
            continue
        target.fill = copy(ref.fill)
        target.border = copy(ref.border)
        target.font = copy(ref.font)
        target.alignment = copy(ref.alignment)


# ── 동적 리뷰 레이아웃 ────────────────────────────────────

def _review_merge_cols(last_data_col: int) -> tuple[int, int]:
    """리뷰 행의 병합 범위 계산.
    Returns (check_end_col, note_start_col)
    - 체크사항: E ~ check_end_col
    - 면책기간: note_start_col ~ L(12)

    핵심: 숨겨진 열(last_data_col+1 ~ K)을 피해서
    면책기간이 항상 보이는 열에서 시작하도록 함.
    """
    if last_data_col >= _DATA_END:
        # 8상품 (열 숨김 없음) → 기존 레이아웃 유지
        return 9, 10  # E:I, J:L

    # 열 숨김 있음 → 체크사항은 E~last_data_col, 면책기간은 L열(12)
    check_end = last_data_col           # 마지막 보이는 데이터 열
    check_end = max(check_end, 6)       # 최소 F(6) — 체크사항 E~F 2열 확보
    note_start = _MAX_COL               # L열(12) — 항상 보이는 열
    return check_end, note_start


def _merge_review_row(ws, r: int, check_end: int, note_start: int):
    """리뷰 데이터 행 병합 생성 (동적 범위)"""
    from openpyxl.utils import get_column_letter
    ws.merge_cells(f"A{r}:B{r}")
    ce = get_column_letter(check_end)
    ws.merge_cells(f"E{r}:{ce}{r}")
    ns = get_column_letter(note_start)
    ws.merge_cells(f"{ns}{r}:L{r}")


def _write_review_header(ws, title_row: int, header_row: int,
                         check_end: int, note_start: int):
    """리뷰 타이틀 + 헤더 행 생성 (동적 병합 범위)"""
    from openpyxl.utils import get_column_letter
    ce = get_column_letter(check_end)
    ns = get_column_letter(note_start)

    ws.merge_cells(f"A{title_row}:L{title_row}")
    safe_val(ws, title_row, 1, "📋  현재 유지중인 보험 리뷰")
    ws.merge_cells(f"A{header_row}:B{header_row}")
    ws.merge_cells(f"E{header_row}:{ce}{header_row}")
    ws.merge_cells(f"{ns}{header_row}:L{header_row}")

    safe_val(ws, header_row, 1, "보험사 / 상품명")
    safe_val(ws, header_row, 3, "가입일 / 만기")
    safe_val(ws, header_row, 4, "월 보험료")
    safe_val(ws, header_row, 5, "주요 체크사항")
    safe_val(ws, header_row, note_start, "특이사항 (면책기간, 보장범위 등)")

    _hdr_font = Font(name=_FONT_NAME, size=9, bold=True, color="FFFFFF")
    _hdr_fill = PatternFill(patternType="solid", fgColor="2E75B6")
    _hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_n in range(1, _MAX_COL + 1):
        cell = ws.cell(row=header_row, column=col_n)
        if cell.__class__.__name__ == "MergedCell":
            continue
        cell.font = _hdr_font
        cell.fill = _hdr_fill
        cell.alignment = _hdr_align
    ws.row_dimensions[title_row].height = 28
    ws.row_dimensions[header_row].height = 20


# ── 내부 유틸 ─────────────────────────────────────────────

def _unmerge_from_row(ws, from_row: int):
    """from_row 이후의 병합 해제 (insert_rows 전 필수)"""
    from openpyxl.utils.cell import range_boundaries
    to_remove = [str(mg) for mg in ws.merged_cells.ranges if mg.min_row >= from_row]
    for r in to_remove:
        ws.merged_cells.ranges = [m for m in ws.merged_cells.ranges if str(m) != r]
        mc, mr, xc, xr = range_boundaries(r)
        for row in range(mr, xr + 1):
            for col in range(mc, xc + 1):
                if (row, col) in ws._cells and ws._cells[(row, col)].__class__.__name__ == "MergedCell":
                    del ws._cells[(row, col)]


def _force_unmerge_range(ws, start_row: int, end_row: int):
    """start_row ~ end_row 범위의 모든 병합을 강제 해제 + ghost MergedCell 제거."""
    from openpyxl.utils.cell import range_boundaries
    to_remove = []
    for mg in ws.merged_cells.ranges:
        if mg.min_row >= start_row and mg.min_row <= end_row:
            to_remove.append(str(mg))
        elif mg.max_row >= start_row and mg.max_row <= end_row:
            to_remove.append(str(mg))
    for r in to_remove:
        try:
            ws.unmerge_cells(r)
        except Exception:
            pass
    # ghost MergedCell 객체 제거
    for row in range(start_row, end_row + 1):
        for col in range(1, _MAX_COL + 1):
            if (row, col) in ws._cells and ws._cells[(row, col)].__class__.__name__ == "MergedCell":
                del ws._cells[(row, col)]
