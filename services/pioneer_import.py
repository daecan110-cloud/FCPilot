"""엑셀 파일에서 개척 매장 일괄 등록 서비스"""
import openpyxl
from io import BytesIO


# 📋 전체 시트 컬럼 매핑 (헤더 행 기준)
EXPECTED_HEADERS = ["순번", "도시", "섹터", "상호명", "업태", "세부업태", "도로명주소"]


def parse_pioneer_excel(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """엑셀 파일을 파싱하여 pioneer_shops insert용 dict 리스트 반환.

    Returns:
        (rows, errors) — rows: insert용 dict 리스트, errors: 파싱 경고 목록
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

    # 📋 전체 시트 찾기
    target_sheet = None
    for name in wb.sheetnames:
        if "전체" in name:
            target_sheet = name
            break
    if not target_sheet:
        return [], ["'전체' 시트를 찾을 수 없습니다."]

    ws = wb[target_sheet]
    rows_out: list[dict] = []
    errors: list[str] = []

    # 헤더 행 찾기 (순번 컬럼이 있는 행)
    header_row = None
    col_map: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        vals = [str(c.value).strip() if c.value else "" for c in row]
        if "순번" in vals and "상호명" in vals:
            header_row = row[0].row
            for i, v in enumerate(vals):
                if v:
                    col_map[v] = i
            break

    if header_row is None:
        return [], ["헤더 행(순번, 상호명 등)을 찾을 수 없습니다."]

    # 데이터 행 파싱
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = list(row)

        shop_name = _cell(vals, col_map, "상호명")
        if not shop_name:
            continue

        city = _cell(vals, col_map, "도시")
        sector = _cell(vals, col_map, "섹터")
        category = _cell(vals, col_map, "업태")
        sub_category = _cell(vals, col_map, "세부업태")
        road_addr = _cell(vals, col_map, "도로명주소")
        jibun_addr = _cell(vals, col_map, "지번주소")
        dong = _cell(vals, col_map, "행정동명")

        address = road_addr or jibun_addr
        memo_parts = [p for p in [city, sector, sub_category, dong] if p]
        memo = " / ".join(memo_parts) if memo_parts else ""

        rows_out.append({
            "shop_name": shop_name,
            "address": address,
            "category": category or "기타",
            "memo": memo,
        })

    wb.close()

    if not rows_out:
        errors.append("파싱된 매장 데이터가 0건입니다.")

    return rows_out, errors


def _cell(vals: list, col_map: dict, key: str) -> str:
    """col_map에서 인덱스를 찾아 값 반환"""
    idx = col_map.get(key)
    if idx is None or idx >= len(vals):
        return ""
    v = vals[idx]
    return str(v).strip() if v is not None else ""


def bulk_insert_shops(supabase, fc_id: str, shops: list[dict],
                      batch_size: int = 50) -> tuple[int, list[str]]:
    """pioneer_shops에 일괄 insert.

    Returns:
        (inserted_count, errors)
    """
    inserted = 0
    errors: list[str] = []

    for i in range(0, len(shops), batch_size):
        batch = shops[i:i + batch_size]
        records = [
            {
                "fc_id": fc_id,
                "shop_name": s["shop_name"],
                "address": s.get("address", ""),
                "category": s.get("category", "기타"),
                "memo": s.get("memo", ""),
            }
            for s in batch
        ]
        try:
            res = supabase.table("pioneer_shops").insert(records).execute()
            inserted += len(res.data) if res.data else len(records)
        except Exception as e:
            errors.append(f"배치 {i // batch_size + 1} 실패: {e}")

    return inserted, errors
